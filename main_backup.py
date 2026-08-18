from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
import sqlite3
from datetime import datetime
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI()

# Dossier où seront enregistrés les factures/justificatifs
JUSTIFICATIFS_DIR = "justificatifs"
import os
os.makedirs(JUSTIFICATIFS_DIR, exist_ok=True)
app.mount("/justificatifs", StaticFiles(directory=JUSTIFICATIFS_DIR), name="justificatifs")

def init_db():
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            utilisateur TEXT NOT NULL,
            type TEXT NOT NULL,
            montant REAL NOT NULL,
            motif TEXT NOT NULL,
            categorie TEXT DEFAULT 'GÉNÉRAL',
            entite TEXT DEFAULT '',
            mode_paiement TEXT DEFAULT 'ESPECES',
            horodatage TEXT NOT NULL,
            cloture INTEGER DEFAULT 0
        )
    """)
    
    # Migration automatique si la base existait déjà sans les nouvelles colonnes
    cursor.execute("PRAGMA table_info(transactions)")
    colonnes = [col[1] for col in cursor.fetchall()]
    if "categorie" not in colonnes:
        cursor.execute("ALTER TABLE transactions ADD COLUMN categorie TEXT DEFAULT 'GÉNÉRAL'")
    if "entite" not in colonnes:
        cursor.execute("ALTER TABLE transactions ADD COLUMN entite TEXT DEFAULT ''")
        
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS justificatifs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transaction_id INTEGER NOT NULL,
            nom_fichier TEXT NOT NULL,
            nom_original TEXT NOT NULL,
            type_fichier TEXT DEFAULT '',
            chemin TEXT NOT NULL,
            date_ajout TEXT NOT NULL,
            FOREIGN KEY (transaction_id) REFERENCES transactions(id)
        )
    """)

    conn.commit()
    conn.close()

init_db()

def ajouter_transaction(utilisateur, type_trans, montant, motif, mode_paiement, categorie="GÉNÉRAL", entite=""):
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    heure = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        """INSERT INTO transactions 
           (utilisateur, type, montant, motif, categorie, entite, mode_paiement, horodatage) 
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (utilisateur, type_trans, montant, motif, categorie, entite, mode_paiement, heure)
    )
    transaction_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return transaction_id

def enregistrer_justificatif(transaction_id, nom_fichier, nom_original, type_fichier, chemin):
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    cursor.execute(
        """INSERT INTO justificatifs
           (transaction_id, nom_fichier, nom_original, type_fichier, chemin, date_ajout)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (
            transaction_id,
            nom_fichier,
            nom_original,
            type_fichier or "",
            chemin,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
    )
    conn.commit()
    conn.close()

def cloturer_caisse():
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    cursor.execute("UPDATE transactions SET cloture = 1 WHERE cloture = 0")
    conn.commit()
    conn.close()

def obtenir_historique_et_solde():
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    cursor.execute(
        """SELECT utilisateur, type, montant, motif, mode_paiement, horodatage, categorie, entite 
           FROM transactions WHERE cloture = 0 ORDER BY id DESC LIMIT 40"""
    )
    rows = cursor.fetchall()
    
    cursor.execute("SELECT mode_paiement, type, SUM(montant) FROM transactions WHERE cloture = 0 GROUP BY mode_paiement, type")
    totaux_raw = cursor.fetchall()
    conn.close()

    totaux = {
        "ESPECES": {"ENTREE": 0.0, "SORTIE": 0.0},
        "BANKILY": {"ENTREE": 0.0, "SORTIE": 0.0},
        "MASRVI": {"ENTREE": 0.0, "SORTIE": 0.0},
        "SEDAD": {"ENTREE": 0.0, "SORTIE": 0.0},
        "VIREMENT": {"ENTREE": 0.0, "SORTIE": 0.0},
        "CHEQUE": {"ENTREE": 0.0, "SORTIE": 0.0}
    }
    
    for mode, type_t, total in totaux_raw:
        if mode in totaux and type_t in totaux[mode]:
            totaux[mode][type_t] = total or 0.0

    solde_especes = totaux["ESPECES"]["ENTREE"] - totaux["ESPECES"]["SORTIE"]
    solde_mobile = (
        (totaux["BANKILY"]["ENTREE"] - totaux["BANKILY"]["SORTIE"]) +
        (totaux["MASRVI"]["ENTREE"] - totaux["MASRVI"]["SORTIE"]) +
        (totaux["SEDAD"]["ENTREE"] - totaux["SEDAD"]["SORTIE"])
    )
    solde_banque = (
        (totaux["VIREMENT"]["ENTREE"] - totaux["VIREMENT"]["SORTIE"]) +
        (totaux["CHEQUE"]["ENTREE"] - totaux["CHEQUE"]["SORTIE"])
    )
    solde_total = solde_especes + solde_mobile + solde_banque

    transactions = [
        {
            "id": r[0],
            "utilisateur": r[1],
            "type": r[2],
            "montant": r[3],
            "motif": r[4],
            "mode": r[5],
            "heure": r[6].split(" ")[1] if " " in r[6] else r[6],
            "categorie": r[7] if len(r) > 7 else "GÉNÉRAL",
            "entite": r[8] if len(r) > 8 else "",
            "justificatifs": justificatifs.get(r[0], [])
        }
        for r in rows
    ]

    return {
        "solde_total": solde_total,
        "solde_especes": solde_especes,
        "solde_mobile": solde_mobile,
        "solde_banque": solde_banque,
        "transactions": transactions
    }

@app.get("/api/data")
async def get_data():
    return JSONResponse(obtenir_historique_et_solde())

@app.post("/api/ajouter")
async def api_ajouter(request: Request):
    """
    Enregistre une entrée/sortie avec ou sans justificatif.
    Utilise toujours multipart/form-data côté navigateur.
    Les fichiers sont traités après création de la transaction.
    """

    try:
        form = await request.form()

        utilisateur = str(form.get("utilisateur", "Inconnu")).strip()
        type_trans = str(form.get("type", "")).strip().upper()
        motif = str(form.get("motif", "")).strip()
        mode_paiement = str(form.get("mode_paiement", "ESPECES")).strip()
        categorie = str(form.get("categorie", "GÉNÉRAL")).strip()
        entite = str(form.get("entite", "")).strip()

        montant_raw = str(form.get("montant", "0")).strip().replace(",", ".")
        montant = float(montant_raw)

        if type_trans not in ("ENTREE", "SORTIE"):
            return JSONResponse(
                {"success": False, "message": "Type d'opération invalide."},
                status_code=400
            )

        if montant <= 0:
            return JSONResponse(
                {"success": False, "message": "Le montant doit être supérieur à 0."},
                status_code=400
            )

        if not motif:
            return JSONResponse(
                {"success": False, "message": "Le motif est obligatoire."},
                status_code=400
            )

        # 1. On crée d'abord la transaction.
        transaction_id = ajouter_transaction(
            utilisateur=utilisateur or "Inconnu",
            type_trans=type_trans,
            montant=montant,
            motif=motif,
            mode_paiement=mode_paiement,
            categorie=categorie or "GÉNÉRAL",
            entite=entite
        )

        # 2. Puis on traite les justificatifs, s'il y en a.
        extensions_autorisees = {
            ".jpg", ".jpeg", ".png", ".webp", ".pdf"
        }

        fichiers_enregistres = 0
        erreurs_fichiers = []

        for fichier in form.getlist("justificatifs"):
            if not isinstance(fichier, UploadFile):
                continue

            if not fichier.filename:
                continue

            nom_original = os.path.basename(fichier.filename)
            extension = os.path.splitext(nom_original)[1].lower()

            if extension not in extensions_autorisees:
                erreurs_fichiers.append(
                    f"{nom_original}: format non autorisé"
                )
                continue

            contenu = await fichier.read()

            # Maximum 10 Mo par fichier
            if len(contenu) > 10 * 1024 * 1024:
                erreurs_fichiers.append(
                    f"{nom_original}: fichier supérieur à 10 Mo"
                )
                continue

            if len(contenu) == 0:
                erreurs_fichiers.append(
                    f"{nom_original}: fichier vide"
                )
                continue

            # Nom sécurisé et unique
            nom_unique = (
                f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
                f"_{transaction_id}_{fichiers_enregistres}"
                f"{extension}"
            )

            chemin_relatif = os.path.join(
                JUSTIFICATIFS_DIR,
                nom_unique
            )

            try:
                with open(chemin_relatif, "wb") as sortie:
                    sortie.write(contenu)

                enregistrer_justificatif(
                    transaction_id=transaction_id,
                    nom_fichier=nom_unique,
                    nom_original=nom_original,
                    type_fichier=fichier.content_type or "",
                    chemin=chemin_relatif
                )

                fichiers_enregistres += 1

            except Exception as fichier_error:
                erreurs_fichiers.append(
                    f"{nom_original}: {str(fichier_error)}"
                )

        # La transaction est considérée comme enregistrée même
        # si aucun justificatif n'a été sélectionné.
        resultat = obtenir_historique_et_solde()
        resultat["success"] = True
        resultat["transaction_id"] = transaction_id
        resultat["justificatifs_enregistres"] = fichiers_enregistres

        if erreurs_fichiers:
            resultat["message"] = (
                "Opération enregistrée, mais certains justificatifs "
                "n'ont pas pu être enregistrés."
            )
            resultat["erreurs_fichiers"] = erreurs_fichiers
        else:
            resultat["message"] = "Opération enregistrée avec succès."

        return JSONResponse(resultat)

    except ValueError:
        return JSONResponse(
            {
                "success": False,
                "message": "Le montant saisi est invalide."
            },
            status_code=400
        )

    except Exception as e:
        return JSONResponse(
            {
                "success": False,
                "message": f"Erreur lors de l'enregistrement : {str(e)}"
            },
            status_code=500
        )

@app.post("/api/cloturer")
async def api_cloturer():
    cloturer_caisse()
    return JSONResponse(obtenir_historique_et_solde())

# --- EXPORTATION EXCEL (.XLSX) ---
@app.get("/api/export/excel")
async def export_excel():
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, horodatage, utilisateur, type, categorie, entite, motif, mode_paiement, montant, cloture 
        FROM transactions ORDER BY id DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal de Caisse"

    # En-têtes
    headers = [
        "ID", "Horodatage", "Agent", "Type Opération", 
        "Catégorie", "Client / Banque / Agent", "Motif", 
        "Mode Règlement", "Montant (MRU)", "Montant (MRO)", "Statut Clôture"
    ]
    ws.append(headers)

    # Style des en-têtes
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Données
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    for r in rows:
        montant_mru = r[8]
        montant_mro = montant_mru * 10
        statut = "Clôturé" if r[9] == 1 else "En cours"
        
        row_data = [
            r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], 
            montant_mru, montant_mro, statut
        ]
        ws.append(row_data)

        # Style visuel par type d'opération
        current_row = ws.max_row
        type_cell = ws.cell(row=current_row, column=4)
        if r[3] == "ENTREE":
            type_cell.font = Font(bold=True, color="166534")
        else:
            type_cell.font = Font(bold=True, color="991B1B")

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col_idx).border = thin_border

    # Ajustement des largeurs de colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Journal_Caisse_Nettoyage_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- EXPORTATION CSV ---
@app.get("/api/export/csv")
async def export_csv():
    conn = sqlite3.connect("caisse.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, horodatage, utilisateur, type, categorie, entite, motif, mode_paiement, montant, cloture 
        FROM transactions ORDER BY id DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow([
        "ID", "Horodatage", "Agent", "Type Operation", 
        "Categorie", "Client / Banque / Agent", "Motif", 
        "Mode Reglement", "Montant (MRU)", "Montant (MRO)", "Statut Cloture"
    ])

    for r in rows:
        writer.writerow([
            r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], 
            r[8], r[8] * 10, "Cloture" if r[9] == 1 else "En cours"
        ])

    output.seek(0)
    filename = f"Journal_Caisse_Nettoyage_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/")
async def get_interface():
    html_code = """
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Gestion de Caisse - Entreprise de Nettoyage</title>
        <script src="https://cdn.tailwindcss.com"></script>
        <style>
            @media print {
                body * { visibility: hidden; }
                #rapport-print, #rapport-print * { visibility: visible; }
                #rapport-print { position: absolute; left: 0; top: 0; width: 100%; }
            }
        </style>
    </head>
    <body class="bg-slate-100 p-2 sm:p-4 font-sans">

        <div class="max-w-md mx-auto bg-white rounded-xl shadow-md p-4 sm:p-5 space-y-4">
            
            <!-- Session Utilisateur -->
            <div class="bg-blue-50 p-2 rounded-lg border border-blue-200 flex justify-between items-center">
                <div>
                    <label class="block text-[10px] font-bold text-blue-800 uppercase">Agent de Caisse :</label>
                    <select id="user-select" class="bg-transparent text-sm font-semibold text-blue-900 outline-none">
                        <option value="Responsable Caisse">Responsable Caisse</option>
                        <option value="Agent Support">Agent Support</option>
                    </select>
                </div>
                <span class="text-xs bg-blue-200 text-blue-800 font-bold px-2 py-1 rounded">MRU / MRO</span>
            </div>

            <!-- Tableau de bord ventilation des soldes -->
            <div class="bg-slate-900 text-white p-4 rounded-xl shadow space-y-3">
                <div class="text-center border-b border-slate-800 pb-2">
                    <span class="text-xs text-slate-400 font-medium uppercase tracking-wider">Solde Total En Caisse</span>
                    <div id="solde-total" class="text-3xl font-black text-green-400">0 MRU</div>
                    <div id="solde-total-mro" class="text-xs text-slate-400">Soit: 0 MRO</div>
                </div>
                
                <div class="grid grid-cols-3 gap-1 text-center text-[11px]">
                    <div class="bg-slate-800 p-2 rounded-lg">
                        <span class="text-slate-400 block">💵 Espèces</span>
                        <span id="solde-especes" class="font-bold text-amber-400 block">0 MRU</span>
                    </div>
                    <div class="bg-slate-800 p-2 rounded-lg">
                        <span class="text-slate-400 block">📱 Mobile</span>
                        <span id="solde-mobile" class="font-bold text-cyan-400 block">0 MRU</span>
                    </div>
                    <div class="bg-slate-800 p-2 rounded-lg">
                        <span class="text-slate-400 block">🏦 Banque</span>
                        <span id="solde-banque" class="font-bold text-purple-400 block">0 MRU</span>
                    </div>
                </div>
            </div>

            <!-- Saisie Transaction -->
            <div class="space-y-2 border-t pt-2">
                <h3 class="text-xs font-bold text-slate-600 uppercase">Nouvelle Opération</h3>
                
                <div class="grid grid-cols-2 gap-2">
                    <div>
                        <label class="block text-[10px] font-bold text-slate-500">Catégorie</label>
                        <select id="categorie" class="w-full p-2 border rounded-lg text-xs bg-slate-50 font-medium">
                            <option value="Fond de Caisse / Capital">💰 Fond de Caisse / Capital</option>
                            <option value="Achat Produits & Matériel">🛒 Achat Produits / Matériel</option>
                            <option value="Transport / Carburant">🚗 Transport / Carburant</option>
                            <option value="Facture Client / Banque">🏦 Facture Client / Banque</option>
                            <option value="Avance sur Salaire">👤 Avance Salaire Employé</option>
                            <option value="Prestation Ponctuelle">🧹 Prestation Ponctuelle</option>
                            <option value="Frais Divers">📦 Frais Divers</option>
                        </select>
                    </div>
                    <div>
                        <label class="block text-[10px] font-bold text-slate-500">Client / Banque / Agent</label>
                        <input id="entite" type="text" placeholder="Ex: BMCI, Agent Diallo" class="w-full p-2 border rounded-lg text-xs">
                    </div>
                </div>

                <div class="grid grid-cols-2 gap-2">
                    <input id="motif" type="text" placeholder="Motif (ex: Achat Javel)" class="p-2 border rounded-lg text-xs">
                    <div>
                        <input id="montant" type="number" oninput="calculerMRO()" placeholder="Montant (MRU)" class="w-full p-2 border rounded-lg text-xs font-bold">
                        <div id="mro-live" class="text-[10px] text-slate-500 font-semibold text-right pr-1">0 MRO</div>
                    </div>
                </div>
                
                <div>
                    <label class="block text-[10px] font-bold text-slate-500">Mode de règlement :</label>
                    <select id="mode-paiement" class="w-full p-2 border rounded-lg text-xs bg-slate-50 font-semibold">
                        <option value="ESPECES">💵 Espèces (Cash)</option>
                        <option value="BANKILY">📱 Bankily</option>
                        <option value="MASRVI">📱 Masrvi</option>
                        <option value="SEDAD">📱 Sedad / BimBank</option>
                        <option value="VIREMENT">🏦 Virement Bancaire</option>
                        <option value="CHEQUE">📜 Chèque Bancaire</option>
                    </select>
                </div>

                <!-- Facture / justificatif -->
                <div class="bg-amber-50 border border-amber-200 rounded-lg p-2 space-y-2">
                    <div class="flex items-center justify-between">
                        <label class="text-[10px] font-bold text-amber-800 uppercase">
                            📎 Facture / Justificatif
                        </label>
                        <span class="text-[9px] text-amber-700">JPG, PNG, WEBP ou PDF • max 10 Mo</span>
                    </div>

                    <input
                        id="justificatifs"
                        type="file"
                        accept="image/*,.pdf"
                        multiple
                        class="w-full text-[10px] bg-white border border-amber-200 rounded-lg p-1.5"
                    >

                    <!-- Sur téléphone, ce bouton ouvre directement l'appareil photo -->
                    <label class="block">
                        <span class="text-[10px] text-amber-800 font-semibold">
                            📷 Scanner / photographier le justificatif
                        </span>
                        <input
                            id="scanner"
                            type="file"
                            accept="image/*"
                            capture="environment"
                            class="w-full text-[10px] mt-1 bg-white border border-amber-200 rounded-lg p-1.5"
                        >
                    </label>

                    <div id="fichiers-selectionnes" class="text-[10px] text-slate-600"></div>
                </div>

                <div class="flex gap-2 pt-1">
                    <button type="button" onclick="envoyer('ENTREE')"
 class="w-1/2 bg-green-600 hover:bg-green-700 text-white font-bold py-2 rounded-lg text-xs shadow">
                        + ENTRÉE (Recette)
                    </button>
                    <button type="button" onclick="envoyer('SORTIE')" class="w-1/2 bg-red-600 hover:bg-red-700 text-white font-bold py-2 rounded-lg text-xs shadow">
                        - SORTIE (Dépense)
                    </button>
                </div>
            </div>

            <!-- Historique -->
            <div class="border-t pt-2">
                <h2 class="text-xs font-bold text-slate-500 uppercase mb-2">Historique des Opérations</h2>
                <div id="historique" class="space-y-2 max-h-52 overflow-y-auto pr-1"></div>
            </div>

            <!-- Exportation & Bilan -->
            <div class="border-t pt-3 space-y-2">
                <div class="grid grid-cols-2 gap-2">
                    <a href="/api/export/excel" download class="bg-emerald-600 hover:bg-emerald-700 text-white text-[11px] font-bold py-2 rounded-lg flex items-center justify-center gap-1 shadow">
                        📊 Export Excel (.xlsx)
                    </a>
                    <a href="/api/export/csv" download class="bg-teal-600 hover:bg-teal-700 text-white text-[11px] font-bold py-2 rounded-lg flex items-center justify-center gap-1 shadow">
                        📁 Export CSV
                    </a>
                </div>
                <button type="button" onclick="imprimerPDF()" class="w-full bg-indigo-600 hover:bg-indigo-700 text-white text-xs font-bold py-2 rounded-lg flex items-center justify-center gap-2">
                    📄 Imprimer le Bilan Journalier PDF
                </button>
                <button type="button" onclick="cloturerJournee()" class="w-full bg-slate-800 hover:bg-slate-900 text-white text-xs font-bold py-2 rounded-lg">
                    📊 Clôturer la Journée & Réinitialiser
                </button>
            </div>
        </div>

        <!-- Zone d'impression PDF -->
        <div id="rapport-print" class="hidden p-6">
            <h1 class="text-2xl font-bold border-b pb-2 mb-2">Rapport Journalier - Société de Nettoyage</h1>
            <p class="text-xs text-gray-600 mb-4">Généré le : <span id="print-date"></span></p>
            <div class="mb-4 bg-gray-50 p-3 rounded border text-sm">
                <p><strong>Solde Total Caisse :</strong> <span id="print-total"></span></p>
                <p><strong>Total Espèces :</strong> <span id="print-especes"></span></p>
                <p><strong>Total Mobile Money :</strong> <span id="print-mobile"></span></p>
                <p><strong>Total Banque (Virement/Chèque) :</strong> <span id="print-banque"></span></p>
            </div>
            <h2 class="font-bold text-sm border-b mb-2">Détail des Transactions</h2>
            <div id="print-table"></div>
        </div>

        <script>
            function calculerMRO() {
                const val = document.getElementById('montant').value;
                document.getElementById('mro-live').innerText = val ? (val * 10).toLocaleString() + ' MRO' : '0 MRO';
            }

            async function chargerDonnees() {
                try {
                    const res = await fetch('/api/data');
                    const data = await res.json();
                    mettreAJourUI(data);
                } catch (e) {
                    console.error("Erreur de chargement", e);
                }
            }

            function mettreAJourUI(data) {
                document.getElementById('solde-total').innerText = data.solde_total.toLocaleString() + ' MRU';
                document.getElementById('solde-total-mro').innerText = 'Soit: ' + (data.solde_total * 10).toLocaleString() + ' MRO';
                document.getElementById('solde-especes').innerText = data.solde_especes.toLocaleString() + ' MRU';
                document.getElementById('solde-mobile').innerText = data.solde_mobile.toLocaleString() + ' MRU';
                document.getElementById('solde-banque').innerText = data.solde_banque.toLocaleString() + ' MRU';
                
                const histDiv = document.getElementById('historique');
                histDiv.innerHTML = '';

                if (!data.transactions || data.transactions.length === 0) {
                    histDiv.innerHTML = '<div class="text-center text-xs text-slate-400 py-3">Aucune opération enregistrée</div>';
                    return;
                }

                data.transactions.forEach(t => {
                    const color = t.type === 'ENTREE' ? 'text-green-700 bg-green-50 border-green-200' : 'text-red-700 bg-red-50 border-red-200';
                    const sign = t.type === 'ENTREE' ? '+' : '-';
                    const entiteText = t.entite ? ` • <span class="font-semibold">${t.entite}</span>` : '';
                    
                    histDiv.innerHTML += `
                        <div class="p-2 rounded-lg border ${color} text-xs space-y-1">
                            <div class="flex justify-between items-center font-bold">
                                <span>[${t.mode}] ${t.motif}</span>
                                <span class="text-sm">${sign}${t.montant.toLocaleString()} MRU</span>
                            </div>
                            <div class="flex justify-between items-center text-[10px] text-slate-500">
                                <span>🏷️ ${t.categorie}${entiteText}</span>
                                <span>${t.utilisateur} à ${t.heure}</span>
                            </div>

                            ${t.justificatifs && t.justificatifs.length ? `
                                <div class="pt-1 border-t border-slate-200 mt-1">
                                    <div class="font-semibold text-[10px] mb-1">📎 Justificatif(s)</div>
                                    <div class="flex flex-wrap gap-1">
                                        ${t.justificatifs.map(j => `
                                            <a href="${j.url}" target="_blank"
                                               class="inline-flex items-center gap-1 px-2 py-1 rounded bg-white border border-slate-200 hover:bg-slate-50 text-[10px] text-blue-700">
                                                ${j.type && j.type.includes('pdf') ? '📄' : '🖼️'} ${escapeHtml(j.nom)}
                                            </a>
                                        `).join('')}
                                    </div>
                                </div>
                            ` : ''}
                        </div>
                    `;
                });
            }

            function escapeHtml(text) {
                const div = document.createElement('div');
                div.textContent = text || '';
                return div.innerHTML;
            }

            function afficherFichiersSelectionnes() {
                const input = document.getElementById('justificatifs');
                const scanner = document.getElementById('scanner');
                const zone = document.getElementById('fichiers-selectionnes');
                const fichiers = [...input.files, ...scanner.files];

                if (!fichiers.length) {
                    zone.innerText = 'Aucun justificatif sélectionné.';
                    return;
                }

                zone.innerHTML = fichiers.map(f =>
                    `📎 ${escapeHtml(f.name)} (${(f.size / 1024 / 1024).toFixed(2)} Mo)`
                ).join('<br>');
            }

            document.getElementById('justificatifs').addEventListener('change', afficherFichiersSelectionnes);
            document.getElementById('scanner').addEventListener('change', afficherFichiersSelectionnes);

            async function envoyer(type) {
                const boutonEntree = document.querySelector(
                    'button[onclick="envoyer(\'ENTREE\')"]'
                );
                const boutonSortie = document.querySelector(
                    'button[onclick="envoyer(\'SORTIE\')"]'
                );

                const currentUser = document.getElementById('user-select').value;
                const motif = document.getElementById('motif').value.trim();
                const montant = document.getElementById('montant').value.trim();
                const modePaiement = document.getElementById('mode-paiement').value;
                const categorie = document.getElementById('categorie').value;
                const entite = document.getElementById('entite').value.trim();

                if (!motif) {
                    alert('Veuillez saisir le motif.');
                    document.getElementById('motif').focus();
                    return;
                }

                if (!montant || parseFloat(montant) <= 0) {
                    alert('Veuillez saisir un montant valide.');
                    document.getElementById('montant').focus();
                    return;
                }

                // Récupération des fichiers depuis les deux champs.
                const inputFichiers = document.getElementById('justificatifs');
                const inputScanner = document.getElementById('scanner');

                const fichiers = [
                    ...Array.from(inputFichiers.files || []),
                    ...Array.from(inputScanner.files || [])
                ];

                // Vérification des tailles avant l'envoi.
                const fichierTropGros = fichiers.find(
                    fichier => fichier.size > 10 * 1024 * 1024
                );

                if (fichierTropGros) {
                    alert(
                        `Le fichier "${fichierTropGros.name}" dépasse la limite de 10 Mo.`
                    );
                    return;
                }

                // Désactive les boutons pendant l'enregistrement
                // pour éviter les doubles clics.
                if (boutonEntree) boutonEntree.disabled = true;
                if (boutonSortie) boutonSortie.disabled = true;

                const texteBouton = type === 'ENTREE'
                    ? '+ ENTRÉE (Enregistrement...)'
                    : '- SORTIE (Enregistrement...)';

                const boutonActuel = type === 'ENTREE'
                    ? boutonEntree
                    : boutonSortie;

                const ancienTexte = boutonActuel
                    ? boutonActuel.innerText
                    : '';

                if (boutonActuel) {
                    boutonActuel.innerText = texteBouton;
                }

                try {
                    const formData = new FormData();

                    formData.append('utilisateur', currentUser);
                    formData.append('type', type);
                    formData.append('montant', String(parseFloat(montant)));
                    formData.append('motif', motif);
                    formData.append('mode_paiement', modePaiement);
                    formData.append('categorie', categorie);
                    formData.append('entite', entite);

                    // Important : chaque fichier porte le même nom de champ.
                    fichiers.forEach(fichier => {
                        formData.append(
                            'justificatifs',
                            fichier,
                            fichier.name
                        );
                    });

                    const res = await fetch('/api/ajouter', {
                        method: 'POST',
                        body: formData
                    });

                    let data = null;

                    try {
                        data = await res.json();
                    } catch (jsonError) {
                        throw new Error(
                            `Réponse invalide du serveur (${res.status}).`
                        );
                    }

                    if (!res.ok || !data.success) {
                        throw new Error(
                            data.message ||
                            `Erreur serveur (${res.status}).`
                        );
                    }

                    // Mise à jour immédiate de l'écran.
                    mettreAJourUI(data);

                    // Réinitialisation des champs.
                    document.getElementById('motif').value = '';
                    document.getElementById('montant').value = '';
                    document.getElementById('entite').value = '';
                    inputFichiers.value = '';
                    inputScanner.value = '';
                    document.getElementById(
                        'fichiers-selectionnes'
                    ).innerText = 'Aucun justificatif sélectionné.';
                    document.getElementById(
                        'mro-live'
                    ).innerText = '0 MRO';

                    if (data.justificatifs_enregistres > 0) {
                        alert(
                            `Opération enregistrée avec ${data.justificatifs_enregistres} justificatif(s).`
                        );
                    } else {
                        alert('Opération enregistrée avec succès.');
                    }

                    if (data.erreurs_fichiers && data.erreurs_fichiers.length) {
                        alert(
                            'Attention :\n' +
                            data.erreurs_fichiers.join('\n')
                        );
                    }

                } catch (e) {
                    console.error('Erreur enregistrement:', e);
                    alert(
                        "Impossible d'enregistrer l'opération.\n\n" +
                        e.message
                    );
                } finally {
                    // Réactive toujours les boutons.
                    if (boutonEntree) boutonEntree.disabled = false;
                    if (boutonSortie) boutonSortie.disabled = false;

                    if (boutonActuel && ancienTexte) {
                        boutonActuel.innerText = ancienTexte;
                    }
                }
            }

            function imprimerPDF() {
                document.getElementById('print-date').innerText = new Date().toLocaleString();
                document.getElementById('print-total').innerText = document.getElementById('solde-total').innerText;
                document.getElementById('print-especes').innerText = document.getElementById('solde-especes').innerText;
                document.getElementById('print-mobile').innerText = document.getElementById('solde-mobile').innerText;
                document.getElementById('print-banque').innerText = document.getElementById('solde-banque').innerText;
                document.getElementById('print-table').innerHTML = document.getElementById('historique').innerHTML;
                
                document.getElementById('rapport-print').classList.remove('hidden');
                window.print();
                document.getElementById('rapport-print').classList.add('hidden');
            }

            async function cloturerJournee() {
                if (confirm("Voulez-vous réinitialiser le solde et clôturer la journée ?")) {
                    const res = await fetch('/api/cloturer', { method: 'POST' });
                    const data = await res.json();
                    mettreAJourUI(data);
                }
            }

            chargerDonnees();
            setInterval(chargerDonnees, 3000);
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_code)