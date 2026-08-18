import csv
import io
import os
import secrets
from datetime import datetime

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import psycopg2
from psycopg2.extras import RealDictCursor

# Imports pour le PDF (ReportLab)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

app = FastAPI()
security = HTTPBasic()

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "Caisse2026!")

def verifier_authentification(credentials: HTTPBasicCredentials = Depends(security)):
    correct_username = secrets.compare_digest(credentials.username, ADMIN_USERNAME)
    correct_password = secrets.compare_digest(credentials.password, ADMIN_PASSWORD)
    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Identifiants incorrects",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username

DATABASE_URL = os.getenv("DATABASE_URL")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JUSTIFICATIFS_DIR = os.path.join(BASE_DIR, "justificatifs")
os.makedirs(JUSTIFICATIFS_DIR, exist_ok=True)

app.mount("/justificatifs", StaticFiles(directory=JUSTIFICATIFS_DIR), name="justificatifs")

def get_db():
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL, sslmode="require")
        return conn, "postgres"
    else:
        import sqlite3
        conn = sqlite3.connect("caisse.db")
        return conn, "sqlite"

def init_db():
    conn, db_type = get_db()
    cursor = conn.cursor()
    
    if db_type == "postgres":
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id SERIAL PRIMARY KEY,
                utilisateur VARCHAR(255) NOT NULL,
                type VARCHAR(50) NOT NULL,
                montant DOUBLE PRECISION NOT NULL,
                motif TEXT NOT NULL,
                categorie VARCHAR(255) DEFAULT 'GÉNÉRAL',
                entite VARCHAR(255) DEFAULT '',
                mode_paiement VARCHAR(50) DEFAULT 'ESPECES',
                horodatage VARCHAR(100) NOT NULL,
                cloture INTEGER DEFAULT 0
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS justificatifs (
                id SERIAL PRIMARY KEY,
                transaction_id INTEGER NOT NULL,
                nom_fichier TEXT NOT NULL,
                nom_original TEXT NOT NULL,
                type_fichier TEXT DEFAULT '',
                chemin TEXT NOT NULL,
                date_ajout TEXT NOT NULL,
                FOREIGN KEY (transaction_id) REFERENCES transactions(id) ON DELETE CASCADE
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                utilisateur TEXT NOT NULL,
                type TEXT NOT NULL,
                montant REAL NOT NULL,
                motif TEXT NOT NULL,
                categorie TEXT DEFAULT 'GÉNÉRAL',
                entite TEXT DEFAULT '',
                mode_paiement TEXT DEFAULT 'ESPECES',
                horodatage TEXT NOT NULL,
                cloture INTEGER DEFAULT 0
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS justificatifs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_id INTEGER NOT NULL,
                nom_fichier TEXT NOT NULL,
                nom_original TEXT NOT NULL,
                type_fichier TEXT DEFAULT '',
                chemin TEXT NOT NULL,
                date_ajout TEXT NOT NULL,
                FOREIGN KEY (transaction_id) REFERENCES transactions(id)
            )
        """)

    conn.commit()
    conn.close()

init_db()

def ajouter_transaction(utilisateur, type_trans, montant, motif, mode_paiement, categorie="GÉNÉRAL", entite=""):
    conn, db_type = get_db()
    cursor = conn.cursor()
    heure = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if db_type == "postgres":
        cursor.execute("""
            INSERT INTO transactions (utilisateur, type, montant, motif, categorie, entite, mode_paiement, horodatage) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (utilisateur, type_trans, montant, motif, categorie, entite, mode_paiement, heure))
        transaction_id = cursor.fetchone()[0]
    else:
        cursor.execute("""
            INSERT INTO transactions (utilisateur, type, montant, motif, categorie, entite, mode_paiement, horodatage) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (utilisateur, type_trans, montant, motif, categorie, entite, mode_paiement, heure))
        transaction_id = cursor.lastrowid

    conn.commit()
    conn.close()
    return transaction_id

def enregistrer_justificatif(transaction_id, nom_fichier, nom_original, type_fichier, chemin):
    conn, db_type = get_db()
    cursor = conn.cursor()
    
    query = """
        INSERT INTO justificatifs (transaction_id, nom_fichier, nom_original, type_fichier, chemin, date_ajout)
        VALUES (%s, %s, %s, %s, %s, %s)
    """ if db_type == "postgres" else """
        INSERT INTO justificatifs (transaction_id, nom_fichier, nom_original, type_fichier, chemin, date_ajout)
        VALUES (?, ?, ?, ?, ?, ?)
    """
    
    cursor.execute(query, (transaction_id, nom_fichier, nom_original, type_fichier or '', chemin, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def cloturer_caisse():
    conn, db_type = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE transactions SET cloture = 1 WHERE cloture = 0")
    conn.commit()
    conn.close()

def obtenir_historique_et_solde():
    conn, db_type = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, utilisateur, type, montant, motif, mode_paiement, horodatage, categorie, entite 
        FROM transactions WHERE cloture = 0 ORDER BY id DESC LIMIT 50
    """)
    rows = cursor.fetchall()

    transaction_ids = [r[0] for r in rows]
    justificatifs = {}
    if transaction_ids:
        if db_type == "postgres":
            cursor.execute("SELECT transaction_id, nom_fichier, nom_original, type_fichier, chemin FROM justificatifs WHERE transaction_id = ANY(%s)", (transaction_ids,))
        else:
            placeholders = ','.join(['?'] * len(transaction_ids))
            cursor.execute(f"SELECT transaction_id, nom_fichier, nom_original, type_fichier, chemin FROM justificatifs WHERE transaction_id IN ({placeholders})", transaction_ids)
            
        for row in cursor.fetchall():
            t_id, nom_f, nom_orig, t_type, path = row
            if t_id not in justificatifs:
                justificatifs[t_id] = []
            
            ext = os.path.splitext(nom_orig)[1].lower()
            if not t_type or t_type == "application/octet-stream":
                if ext == ".pdf":
                    t_type = "application/pdf"
                elif ext in [".jpg", ".jpeg", ".png", ".webp"]:
                    t_type = f"image/{ext.replace('.', '')}"

            justificatifs[t_id].append({
                "nom": nom_orig,
                "url": f"/justificatifs/{nom_f}",
                "type": t_type
            })

    cursor.execute("""
        SELECT mode_paiement, type, SUM(montant) 
        FROM transactions WHERE cloture = 0 
        GROUP BY mode_paiement, type
    """)
    totaux_raw = cursor.fetchall()
    conn.close()

    totaux = {
        'ESPECES': {'ENTREE': 0.0, 'SORTIE': 0.0},
        'BANKILY': {'ENTREE': 0.0, 'SORTIE': 0.0},
        'MASRVI': {'ENTREE': 0.0, 'SORTIE': 0.0},
        'SEDAD': {'ENTREE': 0.0, 'SORTIE': 0.0},
        'VIREMENT': {'ENTREE': 0.0, 'SORTIE': 0.0},
        'CHEQUE': {'ENTREE': 0.0, 'SORTIE': 0.0}
    }

    for mode, type_t, total in totaux_raw:
        if mode in totaux and type_t in totaux[mode]:
            totaux[mode][type_t] = float(total or 0.0)

    solde_especes = totaux['ESPECES']['ENTREE'] - totaux['ESPECES']['SORTIE']
    solde_mobile = (
        (totaux['BANKILY']['ENTREE'] - totaux['BANKILY']['SORTIE']) +
        (totaux['MASRVI']['ENTREE'] - totaux['MASRVI']['SORTIE']) +
        (totaux['SEDAD']['ENTREE'] - totaux['SEDAD']['SORTIE'])
    )
    solde_banque = (
        (totaux['VIREMENT']['ENTREE'] - totaux['VIREMENT']['SORTIE']) +
        (totaux['CHEQUE']['ENTREE'] - totaux['CHEQUE']['SORTIE'])
    )
    solde_total = solde_especes + solde_mobile + solde_banque

    transactions = [
        {
            "id": r[0],
            "utilisateur": r[1],
            "type": r[2],
            "montant": float(r[3]),
            "motif": r[4],
            "mode": r[5],
            "heure": r[6].split(' ')[1] if ' ' in str(r[6]) else str(r[6]),
            "categorie": r[7] if r[7] else "GÉNÉRAL",
            "entite": r[8] if r[8] else "",
            "justificatifs": justificatifs.get(r[0], [])
        }
        for r in rows
    ]

    return {
        "solde_total": solde_total,
        "solde_especes": solde_especes,
        "solde_mobile": solde_mobile,
        "solde_banque": solde_banque,
        "transactions": transactions
    }

@app.get("/api/data", dependencies=[Depends(verifier_authentification)])
async def get_data():
    return JSONResponse(obtenir_historique_et_solde())

@app.post("/api/ajouter", dependencies=[Depends(verifier_authentification)])
async def api_ajouter(request: Request):
    content_type = request.headers.get("content-type", "")
    
    if "multipart/form-data" in content_type:
        form = await request.form()
        
        utilisateur = str(form.get("utilisateur", "Inconnu"))
        type_trans = str(form.get("type", ""))
        montant_val = form.get("montant", 0)
        montant = float(montant_val) if montant_val else 0.0
        motif = str(form.get("motif", ""))
        mode_paiement = str(form.get("mode_paiement", "ESPECES"))
        categorie = str(form.get("categorie", "GÉNÉRAL"))
        entite = str(form.get("entite", ""))

        transaction_id = ajouter_transaction(
            utilisateur=utilisateur,
            type_trans=type_trans,
            montant=montant,
            motif=motif,
            mode_paiement=mode_paiement,
            categorie=categorie,
            entite=entite
        )

        fichiers = form.getlist("justificatifs")
        extensions_autorisees = {".jpg", ".jpeg", ".png", ".webp", ".pdf", ".gif", ".bmp"}

        for fichier in fichiers:
            if not isinstance(fichier, UploadFile) or not fichier.filename:
                continue

            extension = os.path.splitext(fichier.filename)[1].lower()
            if extension not in extensions_autorisees:
                continue

            contenu = await fichier.read()
            if len(contenu) == 0:
                continue

            nom_unique = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{transaction_id}_{os.path.basename(fichier.filename)}"
            chemin_physique = os.path.join(JUSTIFICATIFS_DIR, nom_unique)

            with open(chemin_physique, "wb") as sortie:
                sortie.write(contenu)

            enregistrer_justificatif(
                transaction_id=transaction_id,
                nom_fichier=nom_unique,
                nom_original=os.path.basename(fichier.filename),
                type_fichier=fichier.content_type or "",
                chemin=chemin_physique
            )

        return JSONResponse(obtenir_historique_et_solde())

    data = await request.json()
    transaction_id = ajouter_transaction(
        utilisateur=data.get("utilisateur", "Inconnu"),
        type_trans=data.get("type"),
        montant=float(data.get("montant", 0)),
        motif=data.get("motif", ""),
        mode_paiement=data.get("mode_paiement", "ESPECES"),
        categorie=data.get("categorie", "GÉNÉRAL"),
        entite=data.get("entite", "")
    )
    return JSONResponse(obtenir_historique_et_solde())

@app.post("/api/cloturer", dependencies=[Depends(verifier_authentification)])
async def api_cloturer():
    cloturer_caisse()
    return JSONResponse(obtenir_historique_et_solde())

@app.get("/api/export/excel", dependencies=[Depends(verifier_authentification)])
async def export_excel():
    conn, db_type = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, horodatage, utilisateur, type, categorie, entite, motif, mode_paiement, montant, cloture FROM transactions ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal de Caisse"

    headers = ["ID", "Horodatage", "Agent", "Type Opération", "Catégorie", "Client / Banque / Agent", "Motif", "Mode Règlement", "Montant (MRU)", "Montant (MRO)", "Statut Clôture"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    for r in rows:
        montant_mru = float(r[8])
        montant_mro = montant_mru * 10
        statut = "Clôturé" if r[9] == 1 else "En cours"
        
        row_data = [r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], montant_mru, montant_mro, statut]
        ws.append(row_data)
        
        current_row = ws.max_row
        type_cell = ws.cell(row=current_row, column=4)
        if r[3] == "ENTREE":
            type_cell.font = Font(bold=True, color="166534")
        else:
            type_cell.font = Font(bold=True, color="991B1B")
            
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col_idx).border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Journal_Caisse_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/export/pdf", dependencies=[Depends(verifier_authentification)])
async def export_pdf():
    data_summary = obtenir_historique_et_solde()
    conn, db_type = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, horodatage, utilisateur, type, motif, mode_paiement, montant FROM transactions WHERE cloture = 0 ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#1E293B'),
        alignment=1,
        spaceAfter=15
    )

    story.append(Paragraph("<b>JOURNAL ET RELEVÉ DE CAISSE</b>", title_style))
    story.append(Paragraph(f"<font size=9 color='#64748B'>Date d'impression : {datetime.now().strftime('%d/%m/%Y %H:%M')}</font>", styles['Normal']))
    story.append(Spacer(1, 15))

    # Résumé des Soldes
    solde_data = [
        ["Solde Espèces", "Solde Mobile", "Solde Banque", "SOLDE TOTAL"],
        [
            f"{data_summary['solde_especes']:,} MRU",
            f"{data_summary['solde_mobile']:,} MRU",
            f"{data_summary['solde_banque']:,} MRU",
            f"{data_summary['solde_total']:,} MRU"
        ]
    ]
    solde_table = Table(solde_data, colWidths=[130, 130, 130, 145])
    solde_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#475569')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#16A34A')),
    ]))
    story.append(solde_table)
    story.append(Spacer(1, 20))

    # Table des Transactions
    table_data = [["ID", "Horodatage", "Agent", "Type", "Motif", "Mode", "Montant (MRU)"]]
    for r in rows:
        table_data.append([
            str(r[0]),
            str(r[1]),
            str(r[2]),
            str(r[3]),
            str(r[4]),
            str(r[5]),
            f"{float(r[6]):,}"
        ])

    t_table = Table(table_data, colWidths=[30, 95, 80, 55, 140, 65, 70])
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (6, 0), (6, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ]

    for idx, r in enumerate(rows, start=1):
        if r[3] == "ENTREE":
            t_style.append(('TEXTCOLOR', (3, idx), (3, idx), colors.HexColor('#166534')))
        else:
            t_style.append(('TEXTCOLOR', (3, idx), (3, idx), colors.HexColor('#991B1B')))

    t_table.setStyle(TableStyle(t_style))
    story.append(t_table)

    doc.build(story)
    buffer.seek(0)

    filename = f"Releve_Caisse_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/", response_class=HTMLResponse, dependencies=[Depends(verifier_authentification)])
async def get_interface():
    return r"""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Gestion de Caisse Sécurisée</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-slate-100 p-2 sm:p-4 font-sans">
        <div class="max-w-md mx-auto bg-white rounded-xl shadow-md p-4 space-y-4">
            <div class="bg-blue-50 p-2 rounded-lg border border-blue-200 flex justify-between items-center">
                <div>
                    <label class="block text-[10px] font-bold text-blue-800 uppercase">Agent de Caisse :</label>
                    <select id="user-select" class="bg-transparent text-sm font-semibold text-blue-900 outline-none">
                        <option value="Responsable Caisse">Responsable Caisse</option>
                        <option value="Agent Support">Agent Support</option>
                    </select>
                </div>
                <span class="text-xs bg-green-200 text-green-800 font-bold px-2 py-1 rounded">🔒 SÉCURISÉ</span>
            </div>

            <div class="bg-slate-900 text-white p-4 rounded-xl shadow space-y-3">
                <div class="text-center border-b border-slate-800 pb-2">
                    <span class="text-xs text-slate-400 font-medium uppercase tracking-wider">Solde Total En Caisse</span>
                    <div id="solde-total" class="text-3xl font-black text-green-400">0 MRU</div>
                    <div id="solde-total-mro" class="text-xs text-slate-400">Soit: 0 MRO</div>
                </div>
                <div class="grid grid-cols-3 gap-1 text-center text-[11px]">
                    <div class="bg-slate-800 p-2 rounded-lg"><span class="text-slate-400 block">💵 Espèces</span><span id="solde-especes" class="font-bold text-amber-400 block">0 MRU</span></div>
                    <div class="bg-slate-800 p-2 rounded-lg"><span class="text-slate-400 block">📱 Mobile</span><span id="solde-mobile" class="font-bold text-cyan-400 block">0 MRU</span></div>
                    <div class="bg-slate-800 p-2 rounded-lg"><span class="text-slate-400 block">🏦 Banque</span><span id="solde-banque" class="font-bold text-purple-400 block">0 MRU</span></div>
                </div>
            </div>

            <div class="space-y-2 border-t pt-2">
                <h3 class="text-xs font-bold text-slate-600 uppercase">Nouvelle Opération</h3>
                <div class="grid grid-cols-2 gap-2">
                    <select id="categorie" class="w-full p-2 border rounded-lg text-xs bg-slate-50 font-medium">
                        <option value="Fond de Caisse / Capital">💰 Fond de Caisse / Capital</option>
                        <option value="Achat Produits & Matériel">🛒 Achat Produits / Matériel</option>
                        <option value="Transport / Carburant">🚗 Transport / Carburant</option>
                        <option value="Facture Client / Banque">🏦 Facture Client / Banque</option>
                        <option value="Avance sur Salaire">👤 Avance Salaire Employé</option>
                        <option value="Prestation Ponctuelle">🧹 Prestation Ponctuelle</option>
                        <option value="Frais Divers">📦 Frais Divers</option>
                    </select>
                    <input id="entite" type="text" placeholder="Client / Banque" class="w-full p-2 border rounded-lg text-xs">
                </div>

                <div class="grid grid-cols-2 gap-2">
                    <input id="motif" type="text" placeholder="Motif" class="p-2 border rounded-lg text-xs">
                    <input id="montant" type="number" step="any" placeholder="Montant (MRU)" class="w-full p-2 border rounded-lg text-xs font-bold">
                </div>
                
                <select id="mode-paiement" class="w-full p-2 border rounded-lg text-xs bg-slate-50 font-semibold">
                    <option value="ESPECES">💵 Espèces</option>
                    <option value="BANKILY">📱 Bankily</option>
                    <option value="MASRVI">📱 Masrvi</option>
                    <option value="SEDAD">📱 Sedad</option>
                    <option value="VIREMENT">🏦 Virement</option>
                    <option value="CHEQUE">📜 Chèque</option>
                </select>

                <input id="justificatifs" type="file" accept="image/*,.pdf" multiple class="w-full text-[10px] bg-white border rounded-lg p-1.5">

                <div class="flex gap-2 pt-1">
                    <button type="button" onclick="envoyer('ENTREE')" class="w-1/2 bg-green-600 hover:bg-green-700 text-white font-bold py-2 rounded-lg text-xs shadow transition-colors">+ ENTRÉE</button>
                    <button type="button" onclick="envoyer('SORTIE')" class="w-1/2 bg-red-600 hover:bg-red-700 text-white font-bold py-2 rounded-lg text-xs shadow transition-colors">- SORTIE</button>
                </div>
            </div>

            <div class="border-t pt-2">
                <h2 class="text-xs font-bold text-slate-500 uppercase mb-2">Historique</h2>
                <div id="historique" class="space-y-2 max-h-64 overflow-y-auto"></div>
            </div>

            <div class="border-t pt-3 space-y-2">
                <div class="grid grid-cols-2 gap-2">
                    <a href="/api/export/excel" download class="block text-center bg-emerald-600 hover:bg-emerald-700 text-white text-[11px] font-bold py-2 rounded-lg transition-colors">📊 Export Excel</a>
                    <a href="/api/export/pdf" download class="block text-center bg-rose-600 hover:bg-rose-700 text-white text-[11px] font-bold py-2 rounded-lg transition-colors">📄 Export PDF</a>
                </div>
                <button type="button" onclick="cloturerJournee()" class="w-full bg-slate-800 hover:bg-slate-900 text-white text-xs font-bold py-2 rounded-lg transition-colors">📊 Clôturer la Journée</button>
            </div>
        </div>

        <script>
            async function chargerDonnees() {
                try {
                    const res = await fetch('/api/data');
                    if (res.status === 401) { location.reload(); return; }
                    const data = await res.json();
                    
                    document.getElementById('solde-total').innerText = data.solde_total.toLocaleString() + ' MRU';
                    document.getElementById('solde-total-mro').innerText = 'Soit: ' + (data.solde_total * 10).toLocaleString() + ' MRO';
                    document.getElementById('solde-especes').innerText = data.solde_especes.toLocaleString() + ' MRU';
                    document.getElementById('solde-mobile').innerText = data.solde_mobile.toLocaleString() + ' MRU';
                    document.getElementById('solde-banque').innerText = data.solde_banque.toLocaleString() + ' MRU';
                    
                    const histDiv = document.getElementById('historique');
                    histDiv.innerHTML = '';
                    data.transactions.forEach(t => {
                        const color = t.type === 'ENTREE' ? 'text-green-700 bg-green-50' : 'text-red-700 bg-red-50';
                        histDiv.innerHTML += `
                            <div class="p-2 rounded border ${color} text-xs flex justify-between items-center">
                                <div>
                                    <span class="font-bold">[${t.mode}]</span> ${t.motif}
                                    <span class="text-[10px] text-gray-500 block">${t.heure} - ${t.utilisateur}</span>
                                </div>
                                <span class="font-bold">${t.montant.toLocaleString()} MRU</span>
                            </div>`;
                    });
                } catch (e) {
                    console.error("Erreur de chargement:", e);
                }
            }

            async function envoyer(typeTrans) {
                const montantInput = document.getElementById('montant');
                const motifInput = document.getElementById('motif');
                
                if (!montantInput.value || parseFloat(montantInput.value) <= 0) {
                    alert("Veuillez saisir un montant valide.");
                    return;
                }

                const formData = new FormData();
                formData.append('utilisateur', document.getElementById('user-select').value);
                formData.append('type', typeTrans);
                formData.append('montant', montantInput.value);
                formData.append('motif', motifInput.value || 'Opération ' + typeTrans);
                formData.append('mode_paiement', document.getElementById('mode-paiement').value);
                formData.append('categorie', document.getElementById('categorie').value);
                formData.append('entite', document.getElementById('entite').value);

                const files = document.getElementById('justificatifs').files;
                for (let f of files) {
                    formData.append('justificatifs', f);
                }

                try {
                    const res = await fetch('/api/ajouter', { method: 'POST', body: formData });
                    if (res.ok) {
                        montantInput.value = '';
                        motifInput.value = '';
                        document.getElementById('entite').value = '';
                        document.getElementById('justificatifs').value = '';
                        chargerDonnees();
                    } else {
                        alert("Erreur lors de l'enregistrement de l'opération.");
                    }
                } catch (e) {
                    console.error("Erreur réseau:", e);
                    alert("Impossible de contacter le serveur.");
                }
            }

            async function cloturerJournee() {
                if (confirm("Clôturer la journée ?")) {
                    await fetch('/api/cloturer', { method: 'POST' });
                    chargerDonnees();
                }
            }

            chargerDonnees();
        </script>
    </body>
    </html>
    """